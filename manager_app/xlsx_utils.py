from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

def expired_xlsx(expired):
    """
    Downloads all movies as Excel file with a worksheet for each movie category
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-movies.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Delete the default worksheet
    workbook.remove(workbook.active)

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Define the column titles and widths
    columns = [
        ('Employee Name'),
        ('Staff number'),
        ('Course'),
        ('Training Date'),
        ('Expiry Date')
    ]

    row_num = 1

    worksheet = workbook.create_sheet(
        title='Expired',
    )

    fill = PatternFill(
        start_color='FFFFFFFF',
        end_color='FF000000',
        fill_type=None,
    )
    #assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns,1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = border_bottom
        cell.alignment = centered_alignment
        cell.fill = fill
        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    for exp in expired:
        row_num += 1

        # define data for each cell in the row
        row = [
                str(exp.employee_name),
                exp.staff_number,
                str(exp.course),
                exp.training_date,
                exp.training_expiry
            ]
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def trainingList_xlsx(filtered_training):
    """
    Downloads all movies as Excel file with a worksheet for each movie category
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-training-list.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Delete the default worksheet
    workbook.remove(workbook.active)

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Define the column titles and widths
    columns = [
        ('Employee Name'),
        ('Staff number'),
        ('Course'),
        ('Training Date'),
        ('Expiry Date')
    ]

    row_num = 1

    worksheet = workbook.create_sheet(
        title='Trained',
    )

    fill = PatternFill(
        start_color='FFFFFFFF',
        end_color='FF000000',
        fill_type=None,
    )
    #assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns,1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = border_bottom
        cell.alignment = centered_alignment
        cell.fill = fill
        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    for training in filtered_training:
        row_num += 1

        # define data for each cell in the row
        row = [
                str(training.employee_name),
                training.staff_number,
                str(training.course),
                training.training_date,
                training.training_expiry
            ]
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response

def untrainedlist_xlsx(notTrained_list):
    """
    Downloads all movies as Excel file with a worksheet for each movie category
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-nottrained-list.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Delete the default worksheet
    workbook.remove(workbook.active)

    # Define some styles and formatting that will be later used for cells
    header_font = Font(name='Calibri', bold=True)
    centered_alignment = Alignment(horizontal='center')
    border_bottom = Border(
        bottom=Side(border_style='medium', color='FF000000'),
    )
    wrapped_alignment = Alignment(
        vertical='top',
        wrap_text=True
    )

    # Define the column titles and widths
    columns = [
        ('Employee Name'),
        ('Staff number'),
        ('Course'),
        ('Training Date'),
        ('Expiry Date')
    ]

    row_num = 1

    worksheet = workbook.create_sheet(
        title='Trained',
    )

    fill = PatternFill(
        start_color='FFFFFFFF',
        end_color='FF000000',
        fill_type=None,
    )
    #assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns,1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.border = border_bottom
        cell.alignment = centered_alignment
        cell.fill = fill
        # set column width
        column_letter = get_column_letter(col_num)
        column_dimensions = worksheet.column_dimensions[column_letter]
        column_dimensions.width = 20

    for notTrained in notTrained_list:
        print('reached')
        row_num += 1

        # define data for each cell in the row
        row = [
                str(notTrained.employee_name),
                notTrained.staff_number,
                str(notTrained.course),
                notTrained.training_date,
                notTrained.training_expiry
            ]
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response